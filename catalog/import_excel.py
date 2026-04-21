import openpyxl  # Библиотека для чтения файлов .xlsx
from django.db import IntegrityError  # Ошибка нарушения целостности БД (например, дубли слага)
from .models import Category, Product  # Наши модели
from pytils.translit import slugify  # Библиотека для правильной транслитерации кириллицы

def import_from_excel(excel_file):
    """
    Импорт категорий и товаров из Excel файла.
    Листы: 'category', 'products'
    """
    wb = openpyxl.load_workbook(excel_file)  # Открываем книгу Excel
    categories_created = 0  # Счетчик созданных категорий
    products_created = 0  # Счетчик созданных товаров
    errors = []  # Список для сбора сообщений об ошибках

    # --- Импорт категорий ---
    if 'category' in wb.sheetnames:  # Проверяем наличие листа 'category'
        ws = wb['category']  # Берем этот лист
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2): # Читаем со 2-й строки
            if not row or not row[0]: # Пропускаем пустые строки
                continue
            try:
                name = str(row[0]).strip()  # Имя категории
                description = str(row[1]).strip() if len(row) > 1 and row[1] else ''  # Описание
                is_active = True  # По умолчанию активна
                if len(row) > 2:  # Если есть колонка активности
                    val = str(row[2]).strip().lower()
                    is_active = val not in ('0', 'false', 'нет', 'no') # Проверка на негативные значения

                cat, created = Category.objects.get_or_create(  # Создаем или находим
                    name=name,
                    defaults={
                        'description': description,
                        'is_active': is_active,
                        'slug': slugify(name)  # Генерируем слаг для URL
                    }
                )
                if created:
                    categories_created += 1
            except IntegrityError as e:
                errors.append(f"Строка {row_idx} (category): {e}")

    # --- Импорт товаров ---
    if 'products' in wb.sheetnames:  # Проверяем лист товаров
        ws = wb['products']
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:
                continue
            try:
                name = str(row[0]).strip()  # 1. Название
                category_name = str(row[1]).strip() if len(row) > 1 else ''  # 2. Имя категории
                description = str(row[2]).strip() if len(row) > 2 and row[2] else '' # 3. Описание
                price = float(row[3]) if len(row) > 3 and row[3] else 0  # 4. Цена
                discount = int(row[4]) if len(row) > 4 and row[4] else 0  # 5. Скидка
                is_active = True
                if len(row) > 5:
                    val = str(row[5]).strip().lower()
                    is_active = val not in ('0', 'false', 'нет', 'no')

                try:
                    category = Category.objects.get(name=category_name) # Ищем категорию по имени
                except Category.DoesNotExist: # Если категории из Excel нет в базе
                    errors.append(f"Строка {row_idx} (products): категория '{category_name}' не найдена")
                    continue

                product, created = Product.objects.get_or_create( # Пытаемся создать товар
                    name=name,
                    defaults={
                        'category': category,
                        'description': description,
                        'price': price,
                        'discount_percent': discount,
                        'is_active': is_active,
                        'slug': slugify(name)
                    }
                )
                if created:
                    products_created += 1
                else: # Если товар уже есть, обновляем только цену и скидку
                    product.price = price
                    product.discount_percent = discount
                    product.save()

            except (ValueError, IntegrityError) as e: # Обработка ошибок данных (например, текст вместо цены)
                errors.append(f"Строка {row_idx} (products): {e}")

    # Формируем итоговый отчет
    msg = f"Импорт завершён. Категорий создано: {categories_created}, Товаров создано: {products_created}."
    if errors:
        msg += f" Ошибки: {'; '.join(errors)}"
    return msg